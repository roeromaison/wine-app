"""Excelテンプレートのドロップダウンリストから masters.json を生成する。

テンプレートを更新したら、このスクリプトを流し直せばアプリ側の候補も更新される。

使い方:
    cd backend
    .venv/Scripts/python scripts/build_masters.py "path/to/wine_log.xlsx"

シート名は指定しなければ自動で探す（見出し行に country / region / variety が
並んでいるシートを使う）。「ドロップダウンリスト」「Sheet2」など名前が
ファイルによって違うため、名前ではなく中身で判定している。
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

OUT_PATH = Path(__file__).resolve().parent.parent / "app" / "data" / "masters.json"
REQUIRED_HEADERS = {"country", "region", "variety"}

# Excel側で空セル埋めに使われている値。候補に混ぜたくない。
JUNK_VALUES = {"", "\xa0"}


def find_master_sheet(wb) -> tuple[object, int]:
    """マスタ一覧のシートと見出し行を返す。

    記入例シートにも country / region / variety という見出しが並んでいるため、
    見出しが揃っているだけでは判別できない。候補の中から region の行数が
    最も多いシートを選ぶ（マスタ一覧は661行、記入例は数十行）。
    """
    candidates = []
    for ws in wb.worksheets:
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1
        ):
            labels = {str(c).strip().lower() for c in row if c}
            if REQUIRED_HEADERS <= labels:
                candidates.append((ws, row_idx))
                break

    if not candidates:
        raise SystemExit(
            "country / region / variety の見出しが揃ったシートが見つかりませんでした。"
        )

    def region_count(candidate) -> int:
        ws, header_row = candidate
        rows = list(ws.iter_rows(min_row=header_row, values_only=True))
        header = rows[0]
        for index, label in enumerate(header):
            if label and str(label).strip().lower() == "region":
                return len(extract_column(rows[1:], index))
        return 0

    return max(candidates, key=region_count)


def extract_column(rows: list[tuple], index: int) -> list[str]:
    values = []
    for row in rows:
        value = row[index] if index < len(row) else None
        if value is None:
            continue
        text = str(value).strip()
        # 数式セル（=CHAR(160) など）は候補ではないので落とす
        if text in JUNK_VALUES or text.startswith("="):
            continue
        values.append(text)

    seen: set[str] = set()
    return [v for v in values if not (v in seen or seen.add(v))]


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Excelファイルのパスを指定してください。")

    src = Path(sys.argv[1])
    if not src.exists():
        raise SystemExit(f"ファイルが見つかりません: {src}")

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws, header_row = find_master_sheet(wb)
    rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    wb.close()

    header = rows[0]
    body = rows[1:]

    masters: dict[str, list[str]] = {}
    for index, label in enumerate(header):
        if not label:
            continue
        masters[str(label).strip().lower()] = extract_column(body, index)

    payload = {
        "source": src.name,
        "sheet": ws.title,
        "countries": masters.get("country", []),
        "regions": masters.get("region", []),
        "varieties": masters.get("variety", []),
        "colors": masters.get("color", []),
        "styles": masters.get("style", []),
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    print(f"元ファイル: {src}  シート: {ws.title}")
    for key in ("countries", "regions", "varieties", "colors", "styles"):
        print(f"  {key}: {len(payload[key])} 件")
    print(f"書き出し: {OUT_PATH}")


if __name__ == "__main__":
    main()
